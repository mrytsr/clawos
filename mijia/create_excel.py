#!/usr/bin/env python3
"""
创建小米设备Excel表格
"""

from openpyxl import Workbook
from datetime import datetime

# 已发现的设备列表
devices = [
    {"ip": "192.168.1.2", "did": "26b36f82", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.3", "did": "3dcd4d32", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.4", "did": "2406b079", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.5", "did": "2b41c93d", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.6", "did": "33a207fc", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.7", "did": "461fda45", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.8", "did": "1d75a7f2", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.9", "did": "1da1fbd7", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.12", "did": "2cbac28b", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.13", "did": "1d9e8f94", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.14", "did": "2a0a5a78", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.15", "did": "235a33d3", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.16", "did": "17385cde", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.17", "did": "359417cf", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.18", "did": "2e98ea63", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.19", "did": "21456d1b", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.20", "did": "28d9860d", "name": "未知设备", "model": "", "token": ""},
    {"ip": "192.168.1.28", "did": "32844228", "name": "未知设备", "model": "", "token": ""},
]

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "小米设备"

# 表头
headers = ["序号", "IP地址", "设备ID(DID)", "设备名称", "型号", "Token", "备注"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for row_idx, dev in enumerate(devices, 2):
    ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
    ws.cell(row=row_idx, column=2, value=dev["ip"])    # IP
    ws.cell(row=row_idx, column=3, value=dev["did"])   # DID
    ws.cell(row=row_idx, column=4, value=dev["name"])   # 名称
    ws.cell(row=row_idx, column=5, value=dev["model"])  # 型号
    ws.cell(row=row_idx, column=6, value=dev["token"]) # Token
    ws.cell(row=row_idx, column=7, value="局域网发现")   # 备注

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 15

# 添加说明
ws.cell(row=len(devices) + 3, column=1, value="说明：")
ws.cell(row=len(devices) + 4, column=1, value="1. 此表格包含局域网内发现的小米设备")
ws.cell(row=len(devices) + 5, column=1, value="2. 需要从米家APP或小米云获取Token后才能控制")
ws.cell(row=len(devices) + 6, column=1, value="3. 获取Token后填入表格即可使用miiocli控制")
ws.cell(row=len(devices) + 7, column=1, value=f"4. 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

# 保存
output_file = "/home/tjx/.openclaw/workspace/mijia/小米设备列表.xlsx"
wb.save(output_file)
print(f"已保存: {output_file}")
print(f"共 {len(devices)} 台设备")
