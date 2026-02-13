# 九阳宫交易计划爬虫 - 直接保存 Excel

import asyncio, re, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

EXCEL_FILE = "stocks.xlsx"
JSON_FILE = "stocks.json"

HEADERS = [
    "股票名称", "热度", "作者", "作者ID", 
    "观点", "发布时间", "核心逻辑",
    "点赞", "转发", "评论",
    "URL", "是否强逻辑", "抓取时间"
]

async def crawl(pages=10):
    from browser import AgentBrowser
    browser = AgentBrowser()
    all_data = []
    
    for p in range(1, pages + 1):
        url = f"https://www.jiuyangongshe.com/plan?page={p}"
        await browser.open(url)
        snapshot = await browser.snapshot()
        
        rows = parse_snapshot(snapshot)
        for row in rows:
            all_data.append(row)
        
        print(f"Page {p}: {len(rows)} 条")
        await asyncio.sleep(2)
    
    # 保存
    save_excel(all_data, EXCEL_FILE)
    print(f"\n✅ Excel: {EXCEL_FILE}")
    
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON: {JSON_FILE}")
    
    return all_data

def parse_snapshot(text):
    """解析 snapshot 文本"""
    lines = text.split('\n')
    results = []
    current = None
    
    for line in lines:
        stripped = line.strip()
        
        # 新列表项开始
        if 'listitem:' in stripped:
            if current and current.get('stock_name'):
                results.append(current)
            current = {
                'crawl_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            continue
        
        if current is None:
            continue
        
        # 股票名称: link 后跟股票名 (不在 /u/ 路径中)
        if 'link "' in stripped and '/u/' not in stripped:
            match = re.search(r'link "([^"]+)"', stripped)
            if match:
                name = match.group(1).strip()
                # 排除纯数字序号和特殊文本
                if name and not name.isdigit() and '...' not in name:
                    current['stock_name'] = name
        
        # 热度
        if '🔥' in stripped:
            match = re.search(r'🔥(\d+)', stripped)
            if match:
                current['stock_heat'] = int(match.group(1))
        
        # 作者: link 在 /u/ 路径中
        if '/u/' in stripped:
            match = re.search(r'link "([^"]+)"', stripped)
            if match:
                current['author_name'] = match.group(1).strip()
            u_match = re.search(r'/u/([^"\s]+)', stripped)
            if u_match:
                current['author_id'] = u_match.group(1)
        
        # 时间、观点、内容 (2025 或 2026 年份)
        if any(x in stripped for x in ['2025-', '2026-']):
            # 提取时间
            time_match = re.search(r'(2025-\d{2}-\d{2}|2026-\d{2}-\d{2})', stripped)
            if time_match:
                current['publish_time'] = time_match.group(1)
            
            # 提取观点
            if ' 看好' in stripped:
                current['view'] = '看好'
            elif ' 谨慎' in stripped:
                current['view'] = '谨慎'
            elif '说不清' in stripped:
                current['view'] = '说不清'
            
            # 提取核心逻辑 - 取整行内容
            content = stripped
            # 清理不需要的部分
            for prefix in ['link "', '/url: ']:
                content = content.replace(prefix, '', 1)
            # 提取核心逻辑 (时间后面的内容)
            if '2026-' in content:
                content = content.split('2026-')[-1]
            elif '2025-' in content:
                content = content.split('2025-')[-1]
            # 清理观点前缀
            for view in ['看好', '谨慎', '说不清']:
                if content.startswith(view):
                    content = content[len(view):]
                    break
            current['core_logic'] = content.strip()[:500]
    
    # 保存最后一条
    if current and current.get('stock_name'):
        results.append(current)
    
    return results

def save_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "交易计划"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="07C160", end_color="07C160", fill_type="solid")
    
    # 写表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            key = header_to_key(header)
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 设置列宽
    widths = [15, 8, 15, 35, 10, 18, 80, 8, 8, 8, 50, 12, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    wb.save(filename)

def header_to_key(header):
    return {
        "股票名称": "stock_name",
        "热度": "stock_heat",
        "作者": "author_name",
        "作者ID": "author_id",
        "观点": "view",
        "发布时间": "publish_time",
        "核心逻辑": "core_logic",
        "点赞": "likes",
        "转发": "forwards",
        "评论": "comments",
        "URL": "url",
        "是否强逻辑": "is_strong_logic",
        "抓取时间": "crawl_time"
    }.get(header, header)

if __name__ == "__main__":
    asyncio.run(crawl(10))
